"""Read a testcase-generator xlsx and group rows by (component, endpoint, method).

Reuses the canonical column order from
   .claude/skills/testcase-generator/templates/testcase_columns.yaml
so the two skills stay in lock-step.

Output (stdout, JSON):
    {
      "spec_sheet": "<path>",
      "groups": [
        {"component": "weather", "endpoint_url": "/v1/forecast", "method": "GET",
         "pytest_marker": "@pytest.mark.weather",
         "rows": [<row dict>, ...]}
      ],
      "workflow_rows": [<row dict>, ...]
    }
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

LOG = logging.getLogger(__name__)

_THIS = Path(__file__).resolve()
_GT_TEMPLATE = (
    _THIS.parent.parent.parent / "testcase-generator" / "templates" / "testcase_columns.yaml"
)

# Columns the generator absolutely needs. Missing any of these is a hard error.
_REQUIRED_COLUMNS = (
    "tc_id", "category", "case_type", "priority", "component", "pytest_marker",
    "endpoint_url", "method", "expected_status",
)


class SpecSheetError(Exception):
    """Raised when the input xlsx is not in the expected format."""


def _expected_columns() -> list[str]:
    if not _GT_TEMPLATE.exists():
        raise SpecSheetError(
            f"testcase-generator column template missing at {_GT_TEMPLATE}. "
            "The test-generator skill depends on testcase-generator being installed."
        )
    try:
        data = yaml.safe_load(_GT_TEMPLATE.read_text(encoding="utf-8"))
    except yaml.YAMLError as exc:
        raise SpecSheetError(f"Could not parse {_GT_TEMPLATE}: {exc}") from exc
    return [c["name"] for c in data["columns"]]


def _parse_cell(value: Any) -> Any:
    """Best-effort: JSON-decode strings; otherwise return as-is."""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if s.startswith(("{", "[")):
        try:
            return json.loads(s)
        except json.JSONDecodeError:
            return value
    if s.isdigit():
        return int(s)
    return value


def read(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise SpecSheetError(f"file not found: {path}")
    LOG.info("Reading spec sheet %s", path)
    try:
        wb = load_workbook(str(path), data_only=True)
    except Exception as exc:
        raise SpecSheetError(f"openpyxl could not open {path}: {exc}") from exc

    if "TestCases" not in wb.sheetnames:
        raise SpecSheetError(
            f"{path.name} has no 'TestCases' sheet. "
            f"Expected the testcase-generator format (sheets found: {wb.sheetnames})"
        )
    ws = wb["TestCases"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    expected = _expected_columns()
    extras = [h for h in headers if h not in expected]
    missing_expected = [c for c in expected if c not in headers]
    if missing_expected:
        LOG.warning("Spec sheet missing %d optional column(s): %s",
                    len(missing_expected), missing_expected)
    if extras:
        LOG.info("Spec sheet has %d extra column(s) beyond the template: %s",
                 len(extras), extras)

    # Hard-required columns
    missing_required = [c for c in _REQUIRED_COLUMNS if c not in headers]
    if missing_required:
        raise SpecSheetError(
            f"spec sheet missing required column(s): {missing_required}"
        )

    idx = {h: i for i, h in enumerate(headers) if h is not None}
    rows: list[dict[str, Any]] = []
    skipped_empty = 0
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in raw):
            skipped_empty += 1
            continue
        row = {h: _parse_cell(raw[i]) for h, i in idx.items()}
        rows.append(row)
    LOG.info("Loaded %d rows (skipped %d empty)", len(rows), skipped_empty)

    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    workflow_rows: list[dict[str, Any]] = []
    for row in rows:
        category = (row.get("category") or "").strip()
        if category == "cross_api_workflow":
            workflow_rows.append(row)
            continue
        key = (
            str(row.get("component") or "unknown"),
            str(row.get("endpoint_url") or ""),
            str(row.get("method") or "GET").upper(),
        )
        groups[key].append(row)

    out_groups: list[dict[str, Any]] = []
    for (component, endpoint_url, method), entries in groups.items():
        marker = entries[0].get("pytest_marker") or f"@pytest.mark.{component}"
        out_groups.append({
            "component": component,
            "endpoint_url": endpoint_url,
            "method": method,
            "pytest_marker": marker,
            "summary": (entries[0].get("title") or "").split(":")[0].strip(),
            "rows": entries,
        })
    LOG.info("Grouped into %d endpoint(s) plus %d workflow row(s)",
             len(out_groups), len(workflow_rows))

    return {
        "spec_sheet": str(path),
        "groups": out_groups,
        "workflow_rows": workflow_rows,
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    args = ap.parse_args()
    try:
        result = read(Path(args.path))
    except SpecSheetError as exc:
        print(f"[test-generator] {exc}", file=sys.stderr)
        sys.exit(2)
    json.dump(result, sys.stdout, indent=2, default=str)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    main()
